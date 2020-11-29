from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

class WorksheetPlus(Worksheet):

    def set_row_values(self, row_values, row = 1):
        """Set values to a specific row

        Args:
            row_values (list): List or tuple of values to set in the row
            row (int, optional): row in which the values ​​will be included. Defaults to 1.
        """
        for column, value in enumerate(row_values):
            self.cell(column=column+1, row=row, value=value)

    # TODO Create method to use columns with leter notation
    def set_column_values(self, column_values, column = 1):
        """Set values to a specific column

        Args:
            column_values (list): List or tuple of values to set in the column
            column (int, optional): Columns number (only) in which the values will be included.
            Defaults to 1.
        """
        for row, value in enumerate(column_values):
            self.cell(column=column, row=row+1, value=value)


    def set_values_with_array(self, array):
        """Create a table with an array provided, e.g:
        [["A1","B1","C1"]["A2","B2","C2"]]

        Args:
            array (list): List with a list which has the values of the rows
        """
        for row in array:
            self.append(row)

    def merge_range(self, first_column, last_column, first_row, last_row):
        """Merge a sequence of lines in an determinated range.
        See doc for a usage example.

        Args:
            first_column (int): Column to start the merge of a row
            last_column (int): Column which the merge end
            first_row (int): First row to be merged in the range of the columns
            last_row (int): Last row to be merged in the range of the columns
        """
        for row in range(first_row, last_row + 1):
            self.merge_cells(start_row=row,
                            end_row=row,
                            start_column=first_column,
                            end_column=last_column)

    def set_header(self, array, merge_header_entire_row=True):
        """Set the first lines of the sheet with an array of rows, e.g:
        [["Sales in october 2019"]] or [["jan"],["feb"],["mar"]]
        See doc for more usage example.

        Args:
            array (list): List with a list which has the values of the header rows
            merge_header_entire_row (bool, optional): If True, will merge the rows inserted
                in the header, from first to last column in the sheet. Defaults to True.
        """
        self.__insert_rows_at_top(len(array))
        for index, row in enumerate(array, 1):
            self.set_row_values(row, index)

        if merge_header_entire_row:
            self.merge_cells(start_row=1, end_row=1, start_column=1, end_column=self.max_column)

    def __insert_rows_at_top(self, num_rows):
        """inserts a certain number of rows at the beginning of the sheet

        Args:
            num_rows (int): Number of rows to be inserted
        """
        for _ in range(num_rows):
            self.insert_rows(1)

    def insert_pandas_df(self, dataframe, index=False, header=False):
        """Insert an pandas dataframe (you must have pandas library to use this method).

        Args:
            dataframe (obj, DataFrame): Pandas dataframe to insert
            index (bool, optional): Includes index of the df. Defaults to False.
            header (bool, optional): Includes header of the df. Defaults to False.
        """
        for row in dataframe_to_rows(dataframe, index, header):
            self.append(row)
